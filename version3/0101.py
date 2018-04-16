import openpyxl as px
from datetime import datetime as dt
"""
filepath = "pyxl_test.xlsx"

try:
    wb = px.load_workbook(filepath)
except:
    wb = px.Workbook()


ws = wb.active
ws['A1'].value = "テスト"
ws.cell(row=2,column=3).value = dt.now()

if ws['A1'].value == 'テスト':
    print(1)
else:
    print(0)


wb.save(filepath)
"""



from urllib import request
from bs4 import BeautifulSoup
year = int(input("year  "))
month = int(input("month  "))

url = 'http://www.data.jma.go.jp/obd/stats/etrn/view/daily_s1.php?prec_no=44&block_no=47662&year={}&month={}&day=1&view='.format(year,month)
with request.urlopen(url)as f:
    f = f.read()
    f = f.decode('utf8')


soup = BeautifulSoup(f,'lxml')
#print(soup)
major = soup.select('.mtx')
print(len(major))
print("="*100)
for minor in major:
    day = minor.select('.a_print')
    for d in day:
        print(d.getText(),"日")
    #print(minor)
    minimum = minor.select('.data_0_0')
    date_data = []
    for mn in minimum:
        date_data.append(mn.getText())
    print(date_data)
    print("-"*100)
#for a in soup.find_all('a'):
#print(a.getText())


"""
<tr class="mtx" style="text-align:right;">
    <td style="white-space:nowrap">
        <div class="a_print">
            <a href="hourly_s1.php?prec_no=44&amp;block_no=47662&amp;year=2012&amp;month=09&amp;day=15&amp;view=p1">
                15
            </a>
        </div>
    </td>
    <td class="data_0_0">1014.6</td>
    <td class="data_0_0">1018.7</td>
    <td class="data_0_0">0.0</td>
    <td class="data_0_0">0.0</td>
    <td class="data_0_0">0.0</td>
    <td class="data_0_0">28.3</td>
    <td class="data_0_0">32.3</td>
    <td class="data_0_0">26.2</td>
    <td class="data_0_0">72</td>
    <td class="data_0_0">49</td>
    <td class="data_0_0">3.1</td>
    <td class="data_0_0">6.0</td>
    <td class="data_0_0" style="text-align:center">南東</td>
    <td class="data_0_0">9.2</td>
    <td class="data_0_0" style="text-align:center">南東</td>
    <td class="data_0_0">8.5</td>
    <td class="data_0_0">--</td>
    <td class="data_0_0">--</td>
    <td class="data_0_0" style="text-align:left">晴時々曇</td>
    <td class="data_0_0" style="text-align:left">晴後一時曇</td>
</tr>
"""
