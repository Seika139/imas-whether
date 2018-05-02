import openpyxl as px
from datetime import datetime as dt
from urllib import request
from bs4 import BeautifulSoup
import datetime as dt

class Recorder:
    def __init__(self,place,hidsuke):
        self.place = place
        self.year = hidsuke.year
        self.month = hidsuke.month
        self.day = hidsuke.day
        self.date_array = []
        self.tenki_with_date = []
        self.tenki_dic = {}
        self.filepath = "version3/tenki_record/{}の{}年の天気.xlsx".format(self.place,self.year)
        self.sheetname = "{}月".format(self.month)

        prec_no_box = [34,44,62,82]
        block_no_box = [47590,47662,47772,47807]
        if self.place == "仙台":
            prec_no = prec_no_box[0]
            block_no = block_no_box[0]
        elif self.place == "東京":
            prec_no = prec_no_box[1]
            block_no = block_no_box[1]
        elif self.place == "大阪":
            prec_no = prec_no_box[2]
            block_no = block_no_box[2]
        elif self.place == "福岡":
            prec_no = prec_no_box[3]
            block_no = block_no_box[3]

        self.url = 'http://www.data.jma.go.jp/obd/stats/etrn/view/da'
        self.url += 'ily_s1.php?prec_no={}&block_no={}&year={}&'.format(prec_no,block_no,self.year)
        if self.month <= 9:
            self.url += 'month=0{}&day=1&view='.format(self.month)
        else:
            self.url += 'month={}&day=1&view='.format(self.month)


    def printer(self):
        text = "今日の日付は,"
        text += "{}年{}月{}日 です。".format(self.year,self.month,self.day)
        print(text)

    def print_info(self):
        with request.urlopen(self.url) as f:
            f = f.read()
            f = f.decode('utf8')
        soup = BeautifulSoup(f,'lxml')
        major = soup.select('.mtx')
        print("="*70)
        for minor in major:
            day = minor.select('.a_print')
            if len(day) == 1:
                print(day[0].getText(),"日")
                #print(minor)
                minimum = minor.select('.data_0_0')
                date_data = []
                for mn in minimum:
                    date_data.append(mn.getText())
                print(date_data)
                print("-"*70)
        #for a in soup.find_all('a'):
        #print(a.getText())

    def get_info(self):
        with request.urlopen(self.url) as f:
            f = f.read()
            f = f.decode('utf8')
        soup = BeautifulSoup(f,'lxml')
        major = soup.select('.mtx')
        for minor in major:
            day = minor.select('.a_print')
            if len(day) == 1:
                self.date_array.append(day[0].getText())
                minimum = minor.select('.data_0_0')
                date_data = []
                for mn in minimum:
                    date_data.append(mn.getText())
                self.tenki_with_date.append(date_data)

    def add_to_excel(self):

        def starting_workbook(st):
            matrix = [[1,1,4,1],[1,2,1,3],[3,2,4,2],[3,3,4,3],[1,4,2,6],[3,4,4,4],[3,5,3,6]]
            matrix += [[1,7,2,9],[3,7,4,7],[3,8,4,8],[3,9,4,9],[1,10,2,11],[3,10,4,10],[3,11,4,11]]
            matrix += [[1,12,2,16],[3,12,4,12],[3,13,3,14],[3,15,3,16],[1,17,4,17],[1,18,2,19]]
            matrix += [[1,20,2,21],[3,20,4,20],[3,21,4,21]]

            for i in range(len(matrix)):
                ws.merge_cells(start_row=matrix[i][0],start_column=matrix[i][1],end_row=matrix[i][2],end_column=matrix[i][3])

            st["A1"].value = "日"
            st["B1"].value = "気圧(hPa)"
            st["B2"].value = "現地"
            st["C2"].value = "海面"
            st["B3"].value = "平均"
            st["C3"].value = "平均"
            st["D1"].value = "降水量(mm)"
            st["D3"].value = "合計"
            st["E3"].value = "最大"
            st["E4"].value = "1時間"
            st["F4"].value = "10分間"
            st["G1"].value = "気温(℃)"
            st["G3"].value = "平均"
            st["H3"].value = "最高"
            st["I3"].value = "最低"
            st["J1"].value = "湿度(％)"
            st["J3"].value = "平均"
            st["K3"].value = "最小"
            st["L1"].value = "風向・風速(m/s)"
            st["L3"].value = "平均風速"
            st["M3"].value = "最大風速"
            st["O3"].value = "最大瞬間風速"
            st["M4"].value = "風速"
            st["N4"].value = "風向"
            st["O4"].value = "風速"
            st["P4"].value = "風向"
            st["Q1"].value = "日照時間(h)"
            st["R1"].value = "雪(cm)"
            st["R3"].value = "降雪"
            st["R4"].value = "合計"
            st["S3"].value = "最深積雪"
            st["S4"].value = "値"
            st["T1"].value = "天気概況"
            st["T3"].value = "6〜18時"
            st["U3"].value = "18〜翌6時"

            from openpyxl.styles import Alignment
            for i in range(2,5):
                for j in range(2,22):
                        ws.cell(row=i,column=j).alignment = Alignment(
                wrap_text=False,  # 折り返し改行
                horizontal='center',  # 水平位置
                vertical='bottom'  # 上下位置
                        )

        try:
            wb = px.load_workbook(self.filepath)
            try:
                ws = wb[self.sheetname]
            except:
                ws = wb.create_sheet(self.sheetname)
                starting_workbook(ws)
        except:
            wb = px.Workbook()
            ws = wb.active
            ws.title = self.sheetname
            starting_workbook(ws)

        def check(a,b,c):
            if c == "int":
                if a.value != b:
                    try: a.value = int(b)
                    except: a.value = b
            elif c == "float":
                if a.value != b:
                    try: a.value = float(b)
                    except: a.value = b

        er_ms = ("",0,0,0)
        for day in range(len(self.date_array)):
            for j in range(20):
                try:
                    check(ws.cell(row=day+5,column=1),self.date_array[day],"int")
                    check(ws.cell(row=day+5,column=j+2),self.tenki_with_date[day][j],"float")
                except Exception as e:
                    ms_new = (self.place,self.year,self.month,day+1)
                    if er_ms != ms_new:
                        er_ms = ms_new
                        print(e)
                        print(self.place,self.year,"/",self.month,"/",day+1,"でのエラーです。")
                        print("以下にそのデータを記載します")
                        print("データ長 :",len(self.tenki_with_date[day]))
                        print("データ :",self.tenki_with_date[day])
                        print()
                    continue

        wb.save(self.filepath)

    def eroor_check(self):
        try:
            wb = px.load_workbook(self.filepath)
            try: ws = wb[self.sheetname]
            except: pass
        except: pass
        for day in range(31):
            for column in range(2,22):
                if ws.cell(row=day+5,column=1).value != None:
                    if ws.cell(row=day+5,column=22).value != None:
                        print(self.place,self.year,"/",self.month,"/",day+1,"でのエラーです。")
                text = str(ws.cell(row=day+5,column=column).value)
                if text[-1] == ")" or text[-1] == "]":
                    text = text[:-1]
                    try: ws.cell(row=day+5,column=column).value = float(text)
                    except: ws.cell(row=day+5,column=column).value = text
        wb.save(self.filepath)

    def get_data(self,hidsuke):
        day = hidsuke.day
        wb = px.load_workbook(self.filepath)
        ws = wb[self.sheetname]
        self.tenki_dic["day"] = ws.cell(row=day+4,column=1).value
        self.tenki_dic["平均気圧"] = {}
        self.tenki_dic["平均気圧"]["現地"] = ws.cell(row=day+4,column=2).value
        self.tenki_dic["平均気圧"]["海面"] = ws.cell(row=day+4,column=3).value
        self.tenki_dic["気温"] = {}
        self.tenki_dic["気温"]["平均"] = ws.cell(row=day+4,column=7).value
        self.tenki_dic["気温"]["最高"] = ws.cell(row=day+4,column=8).value
        self.tenki_dic["気温"]["最低"] = ws.cell(row=day+4,column=9).value
        self.tenki_dic["湿度"] = {}
        self.tenki_dic["湿度"]["平均"] = ws.cell(row=day+4,column=10).value
        self.tenki_dic["湿度"]["最小"] = ws.cell(row=day+4,column=11).value
        self.tenki_dic["降水量"] = {}
        self.tenki_dic["降水量"]["合計降水量"] = ws.cell(row=day+4,column=4).value
        self.tenki_dic["降水量"]["最大降水量"] = {}
        self.tenki_dic["降水量"]["最大降水量"]["1h"] = ws.cell(row=day+4,column=5).value
        self.tenki_dic["降水量"]["最大降水量"]["10m"] = ws.cell(row=day+4,column=6).value
        self.tenki_dic["風"] = {}
        self.tenki_dic["風"]["平均"] = ws.cell(row=day+4,column=12).value
        self.tenki_dic["風"]["最大風速"] = {}
        self.tenki_dic["風"]["最大風速"]["風速"] = ws.cell(row=day+4,column=13).value
        self.tenki_dic["風"]["最大風速"]["風向"] = ws.cell(row=day+4,column=14).value
        self.tenki_dic["風"]["最大瞬間風速"] = {}
        self.tenki_dic["風"]["最大瞬間風速"]["風速"] = ws.cell(row=day+4,column=15).value
        self.tenki_dic["風"]["最大瞬間風速"]["風向"] = ws.cell(row=day+4,column=16).value
        self.tenki_dic["日照時間"] = ws.cell(row=day+4,column=17).value
        self.tenki_dic["雪"] = {}
        self.tenki_dic["雪"]["降雪"] = ws.cell(row=day+4,column=18).value
        self.tenki_dic["雪"]["最深積雪"]= ws.cell(row=day+4,column=19).value
        self.tenki_dic["天気概況"] = {}
        self.tenki_dic["天気概況"]["昼"] = ws.cell(row=day+4,column=20).value
        self.tenki_dic["天気概況"]["夜"]= ws.cell(row=day+4,column=21).value
        return self.tenki_dic


if __name__ == '__main__':
    now = dt.datetime.now()
    day = [now.year,now.month,now.day,now.hour,now.minute]
    print("1998年からの去年までの天気予報を更新するのであれば1を")
    print("任意の年の天気予報を更新するのであれば2を")
    print("今月の天気予報を更新するのであれば3を")
    print("過去の天気予報の情報のチェックをするなら4を")
    que = input("終了するならそれ以外のキーを押してください\n>> ")
    if que == "1":
        print("各地の天気予報の記録を更新します")
        for year in range(1998,day[0]):
            for month in range(1,13):
                for place in ["仙台","東京","大阪","福岡"]:
                    record = Recorder(place,[year,month,day[2]])
                    record.get_info()
                    record.add_to_excel()
    if que == "2":
        year = input("情報を更新したい年を半角で入力してください\n>> ")
        for month in range(1,13):
            for place in ["仙台","東京","大阪","福岡"]:
                record = Recorder(place,[int(year),month,day[2]])
                record.get_info()
                record.add_to_excel()
    if que == "3":
        for place in ["仙台","東京","大阪","福岡"]:
            record = Recorder(place,day)
            record.get_info()
            record.add_to_excel()
    if que == "4":
        print("各地の天気予報の記録を確認します")
        for year in range(1998,day[0]):
            for month in range(1,13):
                for place in ["仙台","東京","大阪","福岡"]:
                    record = Recorder(place,[year,month,day[2]])
                    record.eroor_check()
    if que == "5":
        for month in range(1,13):
            for place in ["仙台","東京","大阪","福岡"]:
                record = Recorder(place,[2008,month,day[2]])
                record.eroor_check()
    if que == "6":
        for month in range(1,day[1]+1):
            for place in ["仙台","東京","大阪","福岡"]:
                record = Recorder(place,[day[0],month,day[2]])
                record.get_info()
                record.add_to_excel()
    else: pass
